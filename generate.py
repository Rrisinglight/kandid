#!/usr/bin/env python3
"""
Генерация синтетических данных испытаний лазерного наведения с БПЛА.

Читает кэшированные реальные погодные данные из JSON, генерирует
правдоподобные измерения полётов и сохраняет результат в Excel-файлы
по структуре из methods.md.
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from pathlib import Path

import numpy as np
from numpy.random import Generator
from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).parent
DEFAULT_INPUT = SCRIPT_DIR / "balanced_selection" / "selected_days_2024-2025.json"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"

# ── Координаты контрольных точек (вблизи центра полигона) ────────────
CONTROL_POINTS: dict[str, tuple[float, float]] = {
    "P1": (55.916134, 36.252087),
    "P2": (55.916021, 36.252544),
    "P3": (55.915843, 36.252192),
    "P4": (55.915762, 36.252681),
    "P5": (55.915958, 36.252389),
}

# ── VH-таблица: средний радиус ошибки в сантиметрах ────────────────
# Строки — высоты (м), столбцы — середины ветровых бинов (м/с).
VH_HEIGHTS = np.array([3.0, 5.0, 10.0, 15.0, 20.0])
VH_WIND_MIDS = np.array([1.5, 2.55, 3.25, 4.0, 6.1])
VH_SIGMA_CM = np.array([
    [ 3.0,  4.0,  5.5,  7.0, 15.0],   # H=3
    [ 4.5,  6.5,  8.0,  9.5, 20.0],   # H=5
    [ 7.0,  9.5, 11.0, 14.0, 25.0],   # H=10
    [11.0, 14.5, 18.0, 24.0, 35.0],   # H=15
    [15.0, 19.0, 22.0, 29.0, 45.0],   # H=20
])


@dataclass
class Config:
    seed: int = 42
    heights: list[int] = field(default_factory=lambda: [3, 5, 10, 15, 20])
    points: list[str] = field(default_factory=lambda: ["P1", "P2", "P3", "P4", "P5"])
    repeats: int = 3
    day_bias_scale: float = 0.07
    point_bias_scale: float = 0.04
    r_cv: float = 0.15
    drift_kappa: float = 0.20
    rayleigh_frac: float = 0.08
    sat_scale: float = 0.003
    cloud_scale: float = 0.015
    outlier_prob: float = 0.005
    outlier_scale: float = 1.40
    start_margin_min: int = 40
    end_margin_min: int = 30


# ═══════════════════════════════════════════════════════════════════════
# Интерполяция VH-таблицы
# ═══════════════════════════════════════════════════════════════════════

def _interp1d_clamp(x: float, xs: np.ndarray, ys: np.ndarray) -> tuple[int, int, float]:
    """Найти два соседних узла и вес для линейной интерполяции с зажимом."""
    if x <= xs[0]:
        return 0, 0, 0.0
    if x >= xs[-1]:
        return len(xs) - 1, len(xs) - 1, 0.0
    idx = int(np.searchsorted(xs, x, side="right")) - 1
    idx = min(idx, len(xs) - 2)
    t = (x - xs[idx]) / (xs[idx + 1] - xs[idx])
    return idx, idx + 1, t


def interpolate_mean_r(h: float, v: float) -> float:
    """Билинейная интерполяция VH-таблицы. Возвращает средний R в мм."""
    hi0, hi1, ht = _interp1d_clamp(h, VH_HEIGHTS, VH_SIGMA_CM[:, 0])
    vi0, vi1, vt = _interp1d_clamp(v, VH_WIND_MIDS, VH_SIGMA_CM[0, :])

    c00 = VH_SIGMA_CM[hi0, vi0]
    c01 = VH_SIGMA_CM[hi0, vi1]
    c10 = VH_SIGMA_CM[hi1, vi0]
    c11 = VH_SIGMA_CM[hi1, vi1]

    sigma_cm = (
        c00 * (1 - ht) * (1 - vt)
        + c01 * (1 - ht) * vt
        + c10 * ht * (1 - vt)
        + c11 * ht * vt
    )
    return float(sigma_cm) * 10.0


# ═══════════════════════════════════════════════════════════════════════
# Генератор расписания измерений
# ═══════════════════════════════════════════════════════════════════════

def _parse_hhmm(s: str) -> time:
    parts = s.split(":")
    return time(int(parts[0]), int(parts[1]))


def generate_day_schedule(
    sunrise: str,
    sunset: str,
    n_points: int,
    n_heights: int,
    n_repeats: int,
    rng: Generator,
    cfg: Config,
) -> list[time]:
    """Генерирует временные метки измерений внутри дневного окна."""
    sr = datetime.combine(datetime.min, _parse_hhmm(sunrise))
    ss = datetime.combine(datetime.min, _parse_hhmm(sunset))

    start = sr + timedelta(minutes=cfg.start_margin_min)
    end = ss - timedelta(minutes=cfg.end_margin_min)
    total_seconds = (end - start).total_seconds()
    if total_seconds < 600:
        start = sr + timedelta(minutes=15)
        end = ss - timedelta(minutes=10)
        total_seconds = (end - start).total_seconds()

    n_total = n_points * n_heights * n_repeats
    avg_gap = total_seconds / (n_total + n_points * 2)

    times: list[datetime] = []
    cursor = start + timedelta(seconds=rng.uniform(0, avg_gap * 0.3))

    for _p in range(n_points):
        for _h in range(n_heights):
            for _r in range(n_repeats):
                jitter = rng.uniform(-15, 15)
                times.append(cursor + timedelta(seconds=jitter))
                cursor += timedelta(seconds=rng.uniform(50, 150))
            cursor += timedelta(seconds=rng.uniform(120, 240))
        cursor += timedelta(seconds=rng.uniform(200, 480))

    if len(times) > n_total:
        times = times[:n_total]

    # Масштабирование в доступное окно
    raw_span = (times[-1] - times[0]).total_seconds()
    if raw_span > 0:
        scale = min(1.0, total_seconds * 0.92 / raw_span)
    else:
        scale = 1.0

    result: list[time] = []
    for t in times:
        offset = (t - times[0]).total_seconds() * scale
        actual = start + timedelta(seconds=offset)
        actual = max(start, min(actual, end))
        result.append(actual.time())

    return result


# ═══════════════════════════════════════════════════════════════════════
# Генератор локального ветра
# ═══════════════════════════════════════════════════════════════════════

def _diurnal_wind_factor(time_frac: float) -> float:
    """
    Суточный ход ветра: утро/вечер спокойнее, середина дня — пик.
    time_frac: 0 = sunrise, 1 = sunset.
    """
    return 0.85 + 0.30 * math.sin(math.pi * time_frac)


def generate_local_wind(
    base_wind: float,
    gusts_max: float,
    time_frac: float,
    rng: Generator,
) -> float:
    diurnal = _diurnal_wind_factor(time_frac)
    gust_ratio = gusts_max / max(base_wind, 0.5)
    turbulence_std = base_wind * 0.08 * (gust_ratio / 2.5)
    noise = rng.normal(0, max(turbulence_std, 0.05))
    v = base_wind * diurnal + noise

    if rng.random() < 0.04:
        v += rng.uniform(0.3, gust_ratio * 0.4)

    return round(max(0.1, v), 2)


# ═══════════════════════════════════════════════════════════════════════
# Генератор локальных порывов
# ═══════════════════════════════════════════════════════════════════════

def generate_local_gusts(
    base_wind: float,
    gusts_max: float,
    v_local: float,
    time_frac: float,
    rng: Generator,
) -> float:
    """Порыв для конкретного измерения: масштабируется от локального V
    через дневной gust-ratio, модулируется суточным ходом и турбулентным шумом."""
    day_ratio = gusts_max / max(base_wind, 0.5)
    base_gust = v_local * day_ratio
    diurnal = 0.9 + 0.2 * math.sin(math.pi * time_frac)
    noise = rng.normal(0, 0.6 * day_ratio)
    g = base_gust * diurnal + noise
    return round(max(v_local + 0.3, min(g, gusts_max * 1.15)), 2)


# ═══════════════════════════════════════════════════════════════════════
# Генератор спутников
# ═══════════════════════════════════════════════════════════════════════

def generate_satellites(cloud_cover: float, day_base_sat: int, rng: Generator) -> int:
    cloud_penalty = 0
    noise = rng.integers(-6, 7)
    sat = day_base_sat - cloud_penalty + noise
    return int(np.clip(sat, 16, 32))


# ═══════════════════════════════════════════════════════════════════════
# Генератор локальной облачности
# ═══════════════════════════════════════════════════════════════════════

def generate_local_cloud(base_cloud: float, rng: Generator) -> float:
    signal = 50.0 + 0.25 * (base_cloud - 50.0)
    noise = rng.normal(0, 20.0)
    return round(float(np.clip(signal + noise, 0, 100)), 1)


# ═══════════════════════════════════════════════════════════════════════
# Генератор локального направления ветра
# ═══════════════════════════════════════════════════════════════════════

def generate_local_wind_dir(base_dir: float, rng: Generator) -> float:
    noise = rng.normal(0, 8.0)
    return round((base_dir + noise) % 360, 1)


# ═══════════════════════════════════════════════════════════════════════
# Генератор ошибки (r_x, r_y, R)
# ═══════════════════════════════════════════════════════════════════════

def generate_error(
    h: float,
    v: float,
    gusts: float,
    sat: int,
    cloud: float,
    day_bias: float,
    point_bias: float,
    wind_dir: float,
    rng: Generator,
    cfg: Config,
) -> tuple[float, float, float]:
    mean_r = interpolate_mean_r(h, v)
    mean_r *= max(1.0 + day_bias + point_bias, 0.5)
    mean_r *= 1.0 + cfg.sat_scale * max(0.0, 24.0 - sat)
    mean_r *= 1.0 + cfg.cloud_scale * (cloud / 100.0)
    mean_r = max(mean_r, 0.5)

    if rng.random() < cfg.outlier_prob:
        mean_r *= cfg.outlier_scale

    dir_rad = math.radians(wind_dir)
    v_ref = float(np.median(VH_WIND_MIDS))

    if rng.random() < cfg.rayleigh_frac:
        sigma_axis = mean_r / math.sqrt(math.pi / 2.0)
        drift = sigma_axis * cfg.drift_kappa * (v / v_ref)
        bias_x = drift * math.sin(dir_rad)
        bias_y = drift * math.cos(dir_rad)
        r_x = rng.normal(bias_x, sigma_axis)
        r_y = rng.normal(bias_y, sigma_axis)
        r = math.hypot(r_x, r_y)
        r = max(r, mean_r * 0.40)
        angle = math.atan2(r_x, r_y)
        r_x = r * math.sin(angle)
        r_y = r * math.cos(angle)
    else:
        r_std = max(3.0, mean_r * cfg.r_cv)
        r = rng.normal(mean_r, r_std)
        r = max(r, 0.1)
        kappa = max(0.0, cfg.drift_kappa * v / v_ref)
        angle = rng.vonmises(dir_rad, kappa)
        r_x = r * math.sin(angle)
        r_y = r * math.cos(angle)

    r = max(r, 0.1)

    return round(r_x, 2), round(r_y, 2), round(r, 2)


# ═══════════════════════════════════════════════════════════════════════
# Генерация одного дня
# ═══════════════════════════════════════════════════════════════════════

def generate_day(day: dict, cfg: Config, rng: Generator) -> list[dict]:
    n_pts = len(cfg.points)
    n_hts = len(cfg.heights)
    n_rep = cfg.repeats
    n_total = n_pts * n_hts * n_rep

    schedule = generate_day_schedule(
        day["sunrise"], day["sunset"], n_pts, n_hts, n_rep, rng, cfg,
    )

    day_bias = rng.normal(0, cfg.day_bias_scale)
    day_base_sat = rng.integers(19, 30)

    point_biases = {p: rng.normal(0, cfg.point_bias_scale) for p in cfg.points}

    base_wind: float = day["wind_speed_10m_mean"]
    gusts_max: float = day["wind_gusts_10m_max"]
    base_dir: float = day["wind_direction_10m_mean"]
    base_cloud: float = day["cloud_cover_mean"]

    rows: list[dict] = []
    idx = 0
    for point in cfg.points:
        for height in cfg.heights:
            for rep in range(1, n_rep + 1):
                t = schedule[idx]
                sr = _parse_hhmm(day["sunrise"])
                ss = _parse_hhmm(day["sunset"])
                day_span = (
                    datetime.combine(datetime.min, ss)
                    - datetime.combine(datetime.min, sr)
                ).total_seconds()
                elapsed = (
                    datetime.combine(datetime.min, t)
                    - datetime.combine(datetime.min, sr)
                ).total_seconds()
                time_frac = max(0.0, min(1.0, elapsed / max(day_span, 1)))

                v = generate_local_wind(base_wind, gusts_max, time_frac, rng)
                gusts_local = generate_local_gusts(
                    base_wind, gusts_max, v, time_frac, rng,
                )
                sat = generate_satellites(base_cloud, day_base_sat, rng)
                cloud_local = generate_local_cloud(base_cloud, rng)
                wind_dir_local = generate_local_wind_dir(base_dir, rng)

                r_x, r_y, r = generate_error(
                    height, v, gusts_local, sat, cloud_local,
                    day_bias, point_biases[point], wind_dir_local,
                    rng, cfg,
                )

                rows.append({
                    "time": t.strftime("%H:%M:%S"),
                    "point": point,
                    "H": height,
                    "flight": rep,
                    "V": v,
                    "gusts": gusts_local,
                    "wind_dir": wind_dir_local,
                    "cloud": cloud_local,
                    "sat": sat,
                    "r_x": r_x,
                    "r_y": r_y,
                    "R": r,
                })
                idx += 1

    return rows


# ═══════════════════════════════════════════════════════════════════════
# Запись Excel
# ═══════════════════════════════════════════════════════════════════════

def _format_date_ru(date_str: str) -> str:
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return d.strftime("%d.%m.%Y")


def write_excel(day: dict, rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    sheet_name = _format_date_ru(day["date"])
    ws = wb.active
    ws.title = sheet_name

    # Строка 1: дата, рассвет, закат
    ws["A1"] = sheet_name
    ws["B1"] = day["sunrise"]
    ws["C1"] = day["sunset"]

    # Строки 3-4: сводка погоды
    ws["A3"] = "Температура (°C)"
    ws["B3"] = "Максимальные порывы ветра (м/с)"
    ws["C3"] = "Погода"

    ws["A4"] = day["temperature_2m_mean"]
    ws["B4"] = day["wind_gusts_10m_max"]
    ws["C4"] = day.get("weather_code_text", "")

    # Строки 6-11: координаты контрольных точек
    ws["A6"] = "Точка"
    ws["B6"] = "N (°)"
    ws["C6"] = "E (°)"

    for i, (name, (lat, lon)) in enumerate(CONTROL_POINTS.items(), start=7):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = lat
        ws[f"C{i}"] = lon

    # Строка 13: заголовки
    headers = [
        "Время", "Точка", "H (м)", "Полёт",
        "V (м/с)", "Порывы, откр. источник (м/с)",
        "Направление ветра", "Облачность (%)",
        "Спутники", "r_x (мм)", "r_y (мм)", "R (мм)",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=13, column=col, value=h)

    # Строки 14+: данные
    for i, row in enumerate(rows, start=14):
        ws.cell(row=i, column=1, value=row["time"])
        ws.cell(row=i, column=2, value=row["point"])
        ws.cell(row=i, column=3, value=row["H"])
        ws.cell(row=i, column=4, value=row["flight"])
        ws.cell(row=i, column=5, value=row["V"])
        ws.cell(row=i, column=6, value=row["gusts"])
        ws.cell(row=i, column=7, value=row["wind_dir"])
        ws.cell(row=i, column=8, value=row["cloud"])
        ws.cell(row=i, column=9, value=row["sat"])
        ws.cell(row=i, column=10, value=row["r_x"])
        ws.cell(row=i, column=11, value=row["r_y"])
        ws.cell(row=i, column=12, value=row["R"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Генерация синтетических данных испытаний лазерного наведения с БПЛА",
    )
    p.add_argument("--seed", type=int, default=42)
    p.add_argument(
        "--input-json", default=str(DEFAULT_INPUT),
        help="Путь к JSON с кэшированными погодными данными",
    )
    p.add_argument(
        "--output-dir", default=str(DEFAULT_OUTPUT_DIR),
        help="Папка для Excel-файлов",
    )
    p.add_argument(
        "--single-file", action="store_true",
        help="Один Excel-файл со всеми днями (листы по датам)",
    )
    return p.parse_args()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    args = parse_args()
    cfg = Config(seed=args.seed)
    rng = np.random.default_rng(cfg.seed)

    input_path = Path(args.input_json)
    output_dir = Path(args.output_dir)

    days: list[dict] = json.loads(input_path.read_text(encoding="utf-8"))
    print(f"Загружено дней: {len(days)}")

    if args.single_file:
        output_dir.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        for day in days:
            rows = generate_day(day, cfg, rng)
            sheet_name = _format_date_ru(day["date"])
            ws = wb.create_sheet(title=sheet_name)

            ws["A1"] = sheet_name
            ws["B1"] = day["sunrise"]
            ws["C1"] = day["sunset"]

            ws["A3"] = "Температура (°C)"
            ws["B3"] = "Максимальные порывы ветра (м/с)"
            ws["C3"] = "Погода"
            ws["A4"] = day["temperature_2m_mean"]
            ws["B4"] = day["wind_gusts_10m_max"]
            ws["C4"] = day.get("weather_code_text", "")

            ws["A6"] = "Точка"
            ws["B6"] = "N (°)"
            ws["C6"] = "E (°)"
            for i, (name, (lat, lon)) in enumerate(CONTROL_POINTS.items(), start=7):
                ws[f"A{i}"] = name
                ws[f"B{i}"] = lat
                ws[f"C{i}"] = lon

            headers = [
                "Время", "Точка", "H (м)", "Полёт",
                "V (м/с)", "Порывы, откр. источник (м/с)",
                "Направление ветра", "Облачность (%)",
                "Спутники", "r_x (мм)", "r_y (мм)", "R (мм)",
            ]
            for col, h in enumerate(headers, start=1):
                ws.cell(row=13, column=col, value=h)

            for j, row in enumerate(rows, start=14):
                ws.cell(row=j, column=1, value=row["time"])
                ws.cell(row=j, column=2, value=row["point"])
                ws.cell(row=j, column=3, value=row["H"])
                ws.cell(row=j, column=4, value=row["flight"])
                ws.cell(row=j, column=5, value=row["V"])
                ws.cell(row=j, column=6, value=row["gusts"])
                ws.cell(row=j, column=7, value=row["wind_dir"])
                ws.cell(row=j, column=8, value=row["cloud"])
                ws.cell(row=j, column=9, value=row["sat"])
                ws.cell(row=j, column=10, value=row["r_x"])
                ws.cell(row=j, column=11, value=row["r_y"])
                ws.cell(row=j, column=12, value=row["R"])

            print(f"  {day['date']}: {len(rows)} измерений")

        single_path = output_dir / "all_days.xlsx"
        wb.save(single_path)
        print(f"\nСохранён: {single_path}")
    else:
        for day in days:
            rows = generate_day(day, cfg, rng)
            out_path = output_dir / f"{day['date']}.xlsx"
            write_excel(day, rows, out_path)
            print(f"  {day['date']}: {len(rows)} измерений → {out_path.name}")

        print(f"\nСохранено файлов: {len(days)} в {output_dir}")


if __name__ == "__main__":
    main()
